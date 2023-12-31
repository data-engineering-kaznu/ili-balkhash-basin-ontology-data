from dataflows import Flow, load, dump_to_path, add_metadata, printer, update_resource, unpivot, find_replace,sort_rows
from dataflows.helpers import ResourceMatcher
import csv 
import openpyxl
import os
import pandas as pd


def merge_water_comsumption_files():
    file_path = "archive/GBD_Water_Consumption"
    file_list = os.listdir(file_path)
    outputxlsx = pd.DataFrame()
    for file in file_list:
       df = pd.concat(pd.read_excel( "archive/GBD_Water_Consumption/"+file, sheet_name=None,  skiprows=[0, 2]), ignore_index=True, sort=False)
       outputxlsx = outputxlsx.append( df, ignore_index=True)
    outputxlsx1 = outputxlsx.iloc[:, :3]
    outputxlsx2 = outputxlsx1.dropna(axis='rows')
    outputxlsx2 = outputxlsx2.drop(outputxlsx2[outputxlsx2['Значение'] == '-'].index)
    outputxlsx2 = outputxlsx2.drop(outputxlsx2[outputxlsx2['Значение'] == 'нб'].index)
    outputxlsx2.to_excel("archive/GBD_Water_Consumption/merged_water_consumption.xlsx", index=False)
 
def merge_water_level_files():
    file_path = "archive/GBD_Water_Level_IBB"
    file_list = os.listdir(file_path)
    outputxlsx = pd.DataFrame()
    for file in file_list:
       df = pd.concat(pd.read_excel( "archive/GBD_Water_Level_IBB/"+file, sheet_name=None,  skiprows=[0, 2]), ignore_index=True, sort=False)
       outputxlsx = outputxlsx.append( df, ignore_index=True)
    outputxlsx1 = outputxlsx.iloc[:, :3]
    outputxlsx2 = outputxlsx1.dropna(axis='rows')
    outputxlsx2 = outputxlsx2.drop(outputxlsx2[outputxlsx2['Значение'] == '-'].index)
    outputxlsx2 = outputxlsx2.drop(outputxlsx2[outputxlsx2['Значение'] == 'нб'].index)
    outputxlsx2.to_excel("archive/GBD_Water_Level_IBB/merged_water_level.xlsx", index=False)

def rename_column(from_name, to_name, resources=None):
    def renamer(rows):
        for row in rows:
            yield dict(
                (k if k != from_name else to_name, v) for k, v in row.items()
            )

    def func(package):
        matcher = ResourceMatcher(resources, package.pkg)
        for resource in package.pkg.descriptor["resources"]:
            if matcher.match(resource["name"]):
                for field in resource.get("schema", {}).get("fields", []):
                    if field["name"] == from_name:
                        field["name"] = to_name
        yield package.pkg
        for res in package:
            if matcher.match(res.res.name):
                yield renamer(res)
            else:
                yield res

    return func

def xlsx_to_csv():
    inputExcelFile1 = 'archive/Water_Basins_KZ.xlsx'
    inputExcelFile2 = 'archive/Water Classes.xlsx'
    inputExcelFile3 = 'archive/Water_regulations_KZ.xlsx'
    inputExcelFile4 = 'archive/Population_KZ.xlsx'
    inputExcelFile5 = 'archive/GBD_Water_Consumption/merged_water_consumption.xlsx'
    inputExcelFile6 = 'archive/GBD_Water_Level_IBB/merged_water_level.xlsx'
    
    wb1 = openpyxl.load_workbook(inputExcelFile1)
    wb2 = openpyxl.load_workbook(inputExcelFile2)
    wb3 = openpyxl.load_workbook(inputExcelFile3)
    wb4 = openpyxl.load_workbook(inputExcelFile4)
    wb5 = openpyxl.load_workbook(inputExcelFile5)
    wb6 = openpyxl.load_workbook(inputExcelFile6)
    
    ws_water_basins_main = wb1["Basins_KZ"]
    ws_water_basins_lakes = wb1["Lakes_KZ"]
    ws_water_basins_rivers = wb1["Rivers_KZ"]
    ws_water_basins_water_consumption = wb1["Water_consumption"]
    ws_water_classes_main = wb2["Water_classes"]
    ws_water_classes_objects = wb2["Classes_of_water_objects_KZ"]
    ws_water_classes_quality = wb2["Water_quality_2006"]
    ws_water_regulations = wb3["Hygienic_water_standards"]
    ws_population_main = wb4["Population_KZ"]
    ws_water_consumption= wb5.worksheets[0]
    ws_water_level= wb6.worksheets[0]
    
    OutputCsvFile1 = csv.writer(open("data/water_basins_kz_v1.csv", 'w'), delimiter=",")
    OutputCsvFile2 = csv.writer(open("data/water_basins_lakes_v1.csv", 'w'), delimiter=",")
    OutputCsvFile3 = csv.writer(open("data/water_basins_rivers_v1.csv", 'w'), delimiter=",")
    OutputCsvFile4 = csv.writer(open("data/water_basins_water_comsumption_v1.csv", 'w'), delimiter=",")
    OutputCsvFile5 = csv.writer(open("data/water_classes_v1.csv", 'w'), delimiter=",")
    OutputCsvFile6 = csv.writer(open("data/water_classes_objects_v1.csv", 'w'), delimiter=",")
    OutputCsvFile7 = csv.writer(open("data/water_classes_quality_v1.csv", 'w'), delimiter=",")
    OutputCsvFile8 = csv.writer(open("data/water_regulations_v1.csv", 'w'), delimiter=",")
    OutputCsvFile9 = csv.writer(open("data/population_main_v1.csv", 'w'), delimiter=",")
    OutputCsvFile10 = csv.writer(open("data/water_consumption_v1.csv", 'w'), delimiter=",")
    OutputCsvFile11 = csv.writer(open("data/water_level_v1.csv", 'w'), delimiter=",")
    
    for eachrow in ws_water_basins_main.rows:
        OutputCsvFile1.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_basins_lakes.rows:
        OutputCsvFile2.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_basins_rivers.rows:
        OutputCsvFile3.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_basins_water_consumption.rows:
        OutputCsvFile4.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_classes_main.rows:
        OutputCsvFile5.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_classes_objects.rows:
        OutputCsvFile6.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_classes_quality.rows:
        OutputCsvFile7.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_regulations.rows:
        OutputCsvFile8.writerow([cell.value for cell in eachrow])
    for eachrow in ws_population_main.rows:
        OutputCsvFile9.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_consumption.rows:
        OutputCsvFile10.writerow([cell.value for cell in eachrow])
    for eachrow in ws_water_level.rows:
        OutputCsvFile11.writerow([cell.value for cell in eachrow])
            
def clean_water_basins_kz():
    with open("data/water_basins_kz_v1.csv","r", newline="") as fin, open("data/water_basins_kz_v2.csv","w") as fout:
        writer=csv.writer(fout)
        for row in csv.reader(fin):
            if any(row):
                row[2] = row[2].replace('"','').replace(',','').strip()
                row[5] = row[5].replace('млн','000000').replace(',','').strip()
                writer.writerow(row[1:-1])
def clean_water_basins_lakes_kz():
    with open("data/water_basins_lakes_v1.csv","r", newline="") as fin, open("data/water_basins_lakes_v2.csv","w") as fout:
        writer=csv.writer(fout)
        for row in csv.reader(fin):
            if any(row):
                writer.writerow(row[4:])

def clean_water_basins_rivers_kz():
    with open("data/water_basins_rivers_v1.csv","r", newline="") as fin, open("data/water_basins_rivers_v2.csv","w") as fout:
        writer=csv.writer(fout)
        for row in csv.reader(fin):
            if any(row):
                writer.writerow(row[0:3])

def clean_water_basins_water_comsumption_kz():
    with open("data/water_basins_water_comsumption_v1.csv","r", newline="") as fin, open("data/water_basins_water_comsumption_v2.csv","w") as fout:
        writer=csv.writer(fout)
        for row in csv.reader(fin):
            if any(row):
                writer.writerow(row[2:8])
                
def clean_water_classes():
    with open("data/water_classes_v1.csv","r", newline="") as fin, open("data/water_classes_v2.csv","w") as fout:
        writer=csv.writer(fout)
        for row in csv.reader(fin):
            if any(row):
                row[1] = row[1].replace('"','').strip()
                row[2] = row[2].replace(',','.').strip()
                writer.writerow(row[0:5])
def clean_water_classes_objects():
    with open("data/water_classes_objects_v1.csv","r", newline="") as fin, open("data/water_classes_objects_v2.csv","w") as fout:
        reader= csv.reader(fin)
        writer = csv.writer(fout)
        next(reader, None)  # skip the headers
        for row in reader:
            if any(row):
                row[4] = row[4].replace('до ','').replace(',','.').strip()
                writer.writerow(row[0:5])
                
def clean_water_classes_quality():
    with open("data/water_classes_quality_v1.csv","r", newline="") as fin, open("data/water_classes_quality_v2.csv","w") as fout:
        reader= csv.reader(fin)
        writer = csv.writer(fout)
        next(reader, None)  # skip the headers
        for row in reader:
            if any(row):
                row[0] = row[0].strip()
                row[10] = row[10].replace('"','').strip()
                writer.writerow(row[0:11])  
                              
def clean_water_regulations():
    with open("data/water_regulations_v1.csv","r", newline="") as fin, open("data/water_regulations_v2.csv","w") as fout:
        reader= csv.reader(fin)
        writer = csv.writer(fout)
        next(reader, None)  # skip the headers
        for row in reader:
            if any(row):
                writer.writerow(row[1:6]) 
                
def clean_population_main():
    with open("data/population_main_v1.csv","r", newline="") as fin, open("data/population_main_v2.csv","w") as fout:
        reader= csv.reader(fin)
        writer = csv.writer(fout)
        next(reader, None)  # skip the headers
        next(reader, None)  # skip the headers
        for row in reader:
            if any(row):
                writer.writerow(row) 
                
def ili_balkhas_basin_ontology_process():
    merge_water_comsumption_files()
    merge_water_level_files()
    xlsx_to_csv()
    clean_water_basins_kz()
    clean_water_basins_lakes_kz()
    clean_water_basins_rivers_kz()
    clean_water_basins_water_comsumption_kz()
    clean_water_classes()
    clean_water_classes_objects()
    clean_water_classes_quality()
    clean_water_regulations()
    clean_population_main()
    
    flow = Flow(
        load("data/water_consumption_v1.csv", format='csv', name='water-consumption'),
        rename_column("Код поста", "device_code", "water-consumption"),
        rename_column("Дата", "date", "water-consumption"),
        rename_column("Значение", "value", "water-consumption"),
        update_resource('water-consumption', path='data/water-consumption'),
        
        load("data/water_level_v1.csv", format='csv', name='water-level'),
        rename_column("Код поста", "device_code", "water-level"),
        rename_column("Дата", "date", "water-level"),
        rename_column("Значение", "value", "water-level"),
        update_resource('water-level', path='data/water-level'),
        
        load("data/water_basins_kz_v2.csv", format='csv', name='water-basins-kz'),
        rename_column("Basins_KZ", "basins_kz", "water-basins-kz"),
        rename_column("Square(sq)", "square", "water-basins-kz"),
        rename_column("Water_resources_KZ(cubicmeter)", "water_resources_kz", "water-basins-kz"),
        rename_column("Regions_KZ", "regions_kz", "water-basins-kz"),
        rename_column("Basins_Population", "basins_population", "water-basins-kz"),
        rename_column("Urban_Basins_Population", "urban_basins_population", "water-basins-kz"),
        rename_column("Rural_Basins_Population", "rural_basins_population", "water-basins-kz"),
        rename_column("Rivers_of_Basins", "rivers", "water-basins-kz"),
        rename_column("River_length_in_KZ()", "river_length_in_kz", "water-basins-kz"),
        update_resource('water-basins-kz', path='data/water-basins-kz'),
        
        load("data/water_basins_lakes_v2.csv", format='csv', name='water-basins-lakes'),
        rename_column("Lakes_KZ", "lakes_kz", "water-basins-lakes"),
        rename_column("Square,Â²", "square", "water-basins-lakes"),
        rename_column("Regions", "regions", "water-basins-lakes"),
        update_resource('water-basins-lakes', path='data/water-basins-lakes'),
        
        load("data/water_basins_rivers_v2.csv", format='csv', name='water-basins-rivers'),
        rename_column("Rivers_KZ", "rivers_kz", "water-basins-rivers"),
        rename_column("River_Length", "river_length", "water-basins-rivers"),
        rename_column("River_Length_in_KZ", "river_length_in_kz", "water-basins-rivers"),
        update_resource('water-basins-rivers', path='data/water-basins-rivers'),
        
        load("data/water_basins_water_comsumption_v2.csv", format='csv', name='water-basins-water-comsumption'),
        rename_column("Rivers_KZ", "rivers_kz", "water-basins-water-comsumption"),
        rename_column("Rivers_length,", "rivers_length", "water-basins-water-comsumption"),
        rename_column("River_fall,m", "rivers_fall_m", "water-basins-water-comsumption"),
        rename_column("Average_annual_water_consumption,m3/s", "avg_annual_water_consumption_m3_s", "water-basins-water-comsumption"),
        rename_column("Water_and_energy_resources,Power,thousandkW", "water_and_energy_resources_power_thousand_kw", "water-basins-water-comsumption"),
        rename_column("Waterandenergyresources,Energy,millionkWh/year", "water_and_energy_resources_power_mln_kwh_year", "water-basins-water-comsumption"),        
        update_resource('water-basins-water-comsumption', path='data/water-basins-water-comsumption'),
    
        load("data/water_classes_v2.csv", format='csv', name='water-classes'),
        rename_column("Class", "class", "water-classes"),
        rename_column("Water quality characteristic", "water_quality_characteristics", "water-classes"),
        rename_column("Water pollution index (WPI)", "Water_pollution_index_wpi", "water-classes"),
        rename_column("Domestic and drinking water use", "domestic_and_drinking_water_use", "water-classes"),
        rename_column("Domestic water use", "domestic_water_use", "water-classes"),
        update_resource('water-classes', path='data/water-classes'),

        load("data/water_classes_objects_v2.csv", format='csv', name='water-classes-objects'),
        rename_column("Water objects", "water_objects", "water-classes-objects"),
        rename_column("Type of water objects", "type", "water-classes-objects"),
        rename_column("Region", "region", "water-classes-objects"),
        rename_column("Class", "class", "water-classes-objects"),
        rename_column("WPI", "wpi", "water-classes-objects"),
        sort_rows("wpi","water-classes-objects",reverse=True),
        update_resource('water-classes-objects', path='data/water-classes-objects'),
        update_resource('water-classes-objects', description='Classes and characteristics of water quality according to the value of the complex water pollution index (WPI)'),
    
        load("data/water_classes_quality_v2.csv", format='csv', name='water-classes-quality'),
        rename_column("Rivers_KZ", "rivers", "water-classes-quality"),
        rename_column("Type of water objects", "type", "water-classes-quality"),
        rename_column("Regions", "regions", "water-classes-quality"),
        rename_column("WPI, April 2005", 'wpi_april_2005', "water-classes-quality"),
        rename_column("WPI, March 2006", "wpi_march_2006", "water-classes-quality"),
        rename_column("WPI, April 2006", "wpi_april_2006", "water-classes-quality"),
        rename_column("Ingredients and indicators of water quality", "ingredients_and_indicators", "water-classes-quality"),
        rename_column("Average concentration, mg/l", "avg_concentration_mg/l", "water-classes-quality"),
        rename_column("Multiplicity of exceeding the MPC", "multiplicity_of_exceeding_the_mpc", "water-classes-quality"),
        rename_column("Classes", "classes", "water-classes-quality"),
        rename_column("Water quality characteristic", "characteristic", "water-classes-quality"),
        update_resource('water-classes-quality', path='data/water-classes-quality'),
        update_resource('water-classes-quality', description='State of the quality of surface waters of Kazakhstan in terms of hydrochemical indicators in April 2006'),

        load("data/water_regulations_v2.csv", format='csv', name='water-regulations'),
        rename_column("Substance_name", "substance_name", "water-regulations"),
        rename_column("Indicator type", "indicator_type", "water-regulations"),
        rename_column("Standards (MPC), not more than in mg / l", "standards_mpc_not_more_than_mg/l", "water-regulations"),
        rename_column("Hazard index", "hazard_index", "water-regulations"),
        rename_column("Class", "class", "water-regulations"),
        update_resource('water-regulations', path='data/water-regulations'),
        update_resource('water-regulations', description='Hygienic standards for the content of chemicals in water (to control the migration of harmful chemicals from materials and reagents used in the practice of drinking water supply)'),
        
        load("data/population_main_v2.csv", format='csv', name='population-main'),
        unpivot([
            { 'name': '([0-9]{4})', 'keys': {'year': r'\1'} }
        ], [ {'name': 'year', 'type': 'year'} ], {'name': 'value', 'type': 'string'},resources = "population-main"),
        find_replace(
            fields=[
                {
                    "name": "value",
                    "patterns": [{"find": r"-", "replace": ""}],
                }
            ],
            resources="population-main",
        ),
        rename_column("Regions_KZ", "region", "population-main"),    
        update_resource('population-main', path='data/population-main'),
        
        
        add_metadata(name='ili-balkhash-basin-ontology-data', title='''Ili balkhas basin ontology data'''),
        dump_to_path(),
    )
    flow.process()
    os.remove("data/water_basins_kz_v1.csv")
    os.remove("data/water_basins_kz_v2.csv")
    os.remove("data/water_basins_lakes_v1.csv")
    os.remove("data/water_basins_lakes_v2.csv")
    os.remove("data/water_basins_rivers_v1.csv")
    os.remove("data/water_basins_rivers_v2.csv")
    os.remove("data/water_basins_water_comsumption_v1.csv")
    os.remove("data/water_basins_water_comsumption_v2.csv")
    os.remove("data/water_classes_v1.csv")
    os.remove("data/water_classes_v2.csv")
    os.remove("data/water_classes_objects_v1.csv")
    os.remove("data/water_classes_objects_v2.csv")
    os.remove("data/water_classes_quality_v1.csv")
    os.remove("data/water_classes_quality_v2.csv")
    os.remove("data/water_regulations_v1.csv")
    os.remove("data/water_regulations_v2.csv")
    os.remove("data/population_main_v1.csv")
    os.remove("data/population_main_v2.csv")
    os.remove("data/water_consumption_v1.csv")
    os.remove("data/water_level_v1.csv")
    
if __name__ == '__main__':
    ili_balkhas_basin_ontology_process()
